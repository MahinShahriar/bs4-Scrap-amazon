from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook , load_workbook
import time


# If you don't have a workbook
# wb = Workbook()
# ws = wb.active
# ws.title = 'Data'

# if you have a workbook already
wb = load_workbook('Soup.xlsx')
ws = wb.active

def SaveCodeInFile(r,path):
  with open(path, "w") as f:
    f.write(r)


url = "https://www.amazon.com/s?k=developer+t+shirts&crid=KFZ2PYZFKSEM&sprefix=developer+t+shirt%2Caps%2C581&ref=nb_sb_noss_1"
headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36'}
print('loading......')
time.sleep(5)
print('Fetching data .........')
r = requests.get(url ,headers=headers)
print('data fetched... loading....')
time.sleep(5)
soup = BeautifulSoup(r.text, 'html.parser')
# SaveCodeInFile(url,headers, "data/amzn.html")
print('SuccessFully Fetched  !')
p = soup.prettify()
print(p[0:400])
while True:
  user = input('Are You Wanna save the data? y/n  :')
  if user=='y':
    SaveCodeInFile(p,"data/amzn.html")
    break
  elif user=='n':
    break
    
# vpn = PyVpn(debug=False)
# vpn.start
# print(r.jsonn())
# vpn.stop()
wb.save('Soup.xlsx')